import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def save_to_excel(data, supplier_data, output_file):
    """Сохранение данных в Excel файл, объединяя товары и информацию о продавцах в одном листе."""
    try:
        # Преобразуем данные о товарах в список словарей с плоской структурой
        flat_data = []
        for product in data:
            flat_product = {
                "Артикул": product["article"],
                "Название": product["name"],
                "Бренд": product["brand"],
                "URL": product["url"],
                "Обычная цена": product["price"]["basic"],
                "Цена по ВБ Карте": product["price"]["product"],
                "Цена без ВБ Карты": product["price"]["total"],
                "Отзывы": product["feedbacks"],
                "Рейтинг": product["rating"],
                "Поставщик(продавец)": product["supplier"],
                "ID продавца": product["supplierId"],
                "Рейтинг продавца": product["supplierRating"],
            }
            flat_data.append(flat_product)

        # Создаем DataFrame для товаров
        df = pd.DataFrame(flat_data)

        # Преобразуем данные о продавцах в DataFrame
        flat_supplier_data = [
            {
                "ID продавца": supplier["supplierId"],
                "Название продавца": supplier["supplierName"],
                "Полное юридическое название": supplier["supplierFullName"],
                "ИНН": supplier["inn"],
                "ОГРН": supplier["ogrn"],
                "ОГРНИП": supplier["ogrnip"],
                "Юридический адрес": supplier["legalAddress"],
                "Торговая марка": supplier["trademark"],
                "КПП": supplier["kpp"],
                "Номер регистрации": supplier["taxpayerCode"],
                "УНП": supplier["unp"],
                "БИН": supplier["bin"],
                "УНН": supplier["unn"],
                "Ссылка на продавца": supplier["supplierUrl"],
            }
            for supplier in supplier_data
        ]
        supplier_df = pd.DataFrame(flat_supplier_data)

        # Объединяем данные о товарах и продавцах по ID продавца (left join)
        merged_df = pd.merge(
            df, supplier_df, on="ID продавца", how="left", suffixes=("", "_продавец")
        )

        # Сохраняем в Excel в один лист
        with pd.ExcelWriter(output_file, engine="openpyxl", mode="w") as writer:
            merged_df.to_excel(writer, sheet_name="Товары и продавцы", index=False)

        # Открываем Excel-файл для настройки ширины столбцов
        workbook = load_workbook(output_file)
        worksheet = workbook["Товары и продавцы"]

        # Настройка ширины столбцов
        columns_to_widen_one = [
            "Артикул",
            "Бренд",
            "Обычная цена",
            "Цена по ВБ Карте",
            "Цена без ВБ Карты",
            "ID продавца",
            "ИНН",
            "ОГРН",
            "ОГРНИП",
            "КПП",
            "Номер регистрации",
            "УНП",
            "БИН",
            "УНН",
        ]
        columns_to_widen_two = [
            "Поставщик(продавец)",
            "Название продавца",
            "Торговая марка",
        ]
        columns_to_widen_three = [
            "Название",
            "URL",
            "Полное юридическое название",
            "Юридический адрес",
            "Ссылка на продавца",
        ]
        default_width = 8.43  # Стандартная ширина столбца в openpyxl
        first_width = float(default_width * 1.3)  # Увеличиваем ширину в 1.3 раза
        second_width = float(default_width * 3)  # Увеличиваем ширину в 3 раза
        third_width = float(default_width * 6)  # Увеличиваем ширину в 6 раз

        for col_idx, col_name in enumerate(merged_df.columns, start=1):
            column_letter = get_column_letter(col_idx)
            if col_name in columns_to_widen_one:
                worksheet.column_dimensions[column_letter].width = first_width
            elif col_name in columns_to_widen_two:
                worksheet.column_dimensions[column_letter].width = second_width
            elif col_name in columns_to_widen_three:
                worksheet.column_dimensions[column_letter].width = third_width

        # Сохраняем изменения
        workbook.save(output_file)
        print(f"Создан файл {output_file}")

    except Exception as e:
        raise Exception(f"Ошибка при записи Excel файла {output_file}: {str(e)}")
